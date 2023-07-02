import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill

class WalletInfo:
    def __init__(self, data_file, wallets_file):
        self.data_file = data_file
        self.wallets_file = wallets_file
        self.df = None
        self.wallets = None
        self.wallet_index_map = None
        self.wb = None
        self.ws = None

    def load_data(self):
        self.df = pd.read_csv(self.data_file)

    def load_wallets(self):
        with open(self.wallets_file, 'r') as f:
            self.wallets = [line.strip().lower() for line in f]

    def map_wallets(self):
        self.wallet_index_map = {wallet: idx + 1 for idx, wallet in enumerate(self.wallets)}

    def add_wallet_number_column(self):
        self.df.insert(0, 'wallet_number', self.df['ua'].map(self.wallet_index_map))

    def convert_and_add_column(self):
        self.df['atad'] = self.df['dwm'].str.split(' / ').str[0].astype(int) / self.df['lzd'].astype(int)
        self.df['average'] = self.df['amt'].astype(int) / self.df['tc'].astype(int)
    
    def create_workbook(self):
        self.wb = Workbook()
        self.ws = self.wb.active

    def add_headers(self):
        for i, column_name in enumerate(self.df.columns.tolist(), start=1):
            self.ws.cell(row=1, column=i, value=column_name)

    def process_wallets(self):
        for idx, wallet in enumerate(self.wallets, start=2):
            if wallet in self.df['ua'].values:
                row = self.df[self.df['ua'] == wallet]

                for i, value in enumerate(row.values[0], start=1):
                    self.ws.cell(row=idx, column=i, value=value)

    def save_workbook(self, file_name):
        self.wb.save(file_name)
        return self.wb


class FindSimilarities:
    def __init__(self, data_file):
        self.data_file = data_file
        self.similarities_sheet = None
        self.similarities_3plus_sheet = None
        self.similarities_4plus_sheet = None

    def extract_month_day(self, date_str):
        return date_str.split()[0].split('-')[1:]

    def find_similarities(self):
        wb = load_workbook(self.data_file)
        self.similarities_sheet = wb.create_sheet("similarities")

        # Define the light blue color for the cell
        light_blue_fill = PatternFill(start_color="ADD8E6",
                                    end_color="ADD8E6",
                                    fill_type="solid")
        # Define the light green color for the cell
        light_green_fill = PatternFill(start_color="90EE90",
                                    end_color="90EE90",
                                    fill_type="solid")

        df = pd.read_excel(self.data_file)
        df = df.drop(columns=['rs'])  # Drop the 'rs' column

        last_row_was_blank = False

        for i in range(len(df)):
            added_similar_lines = set()
            main_row = None
            ibt_main = None
            lbt_main = None
            for j in range(i + 1, len(df)):
                row1 = df.iloc[i]
                row2 = df.iloc[j]
                similarities = sum(row1 == row2)

                ibt_row1 = self.extract_month_day(row1['ibt'])
                ibt_row2 = self.extract_month_day(row2['ibt'])
                lbt_row1 = self.extract_month_day(row1['lbt'])
                lbt_row2 = self.extract_month_day(row2['lbt'])

                if similarities >= 4:
                    if i not in added_similar_lines:
                        self.similarities_sheet.append(list(row1))
                        # Color the first line
                        for cell in self.similarities_sheet[str(self.similarities_sheet.max_row)]:
                            cell.fill = light_blue_fill
                        added_similar_lines.add(i)
                        last_row_was_blank = False
                        main_row = list(row1)
                        ibt_main = ibt_row1
                        lbt_main = lbt_row1
                    self.similarities_sheet.append(list(row2))
                    # Color the cells that are similar to the main row or have the same 'ibt' or 'lbt' month-day values
                    if main_row is not None:
                        for idx, (cell_row_main, cell_row2) in enumerate(zip(main_row, self.similarities_sheet[str(self.similarities_sheet.max_row)])):
                            if cell_row_main == cell_row2.value or ((idx == df.columns.get_loc('ibt') and ibt_row2 == ibt_main) or (idx == df.columns.get_loc('lbt') and lbt_row2 == lbt_main)):
                                cell_row2.fill = light_green_fill
                    last_row_was_blank = False

            # Append blank row after checking each line only if the last row isn't already blank
            if not last_row_was_blank:
                self.similarities_sheet.append([])
                last_row_was_blank = True

        wb.save(self.data_file)



def main():
    wallet_info = WalletInfo('lz.csv', 'wallets.txt')
    wallet_info.load_data()
    wallet_info.load_wallets()
    wallet_info.map_wallets()
    wallet_info.add_wallet_number_column()
    wallet_info.convert_and_add_column()
    wallet_info.create_workbook()
    wallet_info.add_headers()
    wallet_info.process_wallets()
    wallet_info.save_workbook('wallet_info.xlsx')

    similarities = FindSimilarities('wallet_info.xlsx')
    similarities.find_similarities()
    print("DONE")

if __name__ == '__main__':
    main()